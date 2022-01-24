import csv, requests, functools, os
import openpyxl as opl
import matplotlib.pyplot as plt

# from bs4 import BeautifulSoup 
from openpyxl.styles import Alignment, Font, Border, Side

""" Criando um novo arquivo xlsx """
arquivo = opl.Workbook()
planilha = arquivo.active

# convertendo o csv em xlsx
with open('Dados-covid-19-municipios.csv') as file:
    reader = csv.reader(file, delimiter=';')
    for row in reader: 
        planilha.append(row) # preenchendo o xlsx com os dados do csv


""" Alterando a coluna de Grande Região para DRS """
planilha.title = 'PLANILHA_MUNICIPIOS'

# alterando o texto da célula que contém 'Grande Região' para 'Nome Da DRS'
planilha["B1"].value = 'Nome Da DRS'

# declarando arrays das DRSes
drsRioPreto = ['ADOLFO', 'ÁLVARES FLORENCE', 'AMÉRICO DE CAMPOS', "APARECIDA D'OESTE", 'ARIRANHA', 'ASPÁSIA', 'BADY BASSITT', 'BÁLSAMO', 
'CARDOSO', 'CATANDUVA', 'CATIGUÁ', 'CEDRAL', 'COSMORAMA', 'DIRCE REIS', 'DOLCINÓPOLIS', 'ELISIÁRIO', 'EMBAÚBA', "ESTRELA D'OESTE", 'FERNANDÓPOLIS', 'FERNANDO PRESTES', 'FLOREAL', 'GASTÃO VIDIGAL', 'GENERAL SALGADO', 'GUAPIAÇU', "GUARANI D'OESTE", 'IBIRÁ', 'ICÉM', 'INDIAPORÃ', 'IPIGUÁ', 'IRAPUÃ', 'ITAJOBI', 'JACI', 'JALES', 'JOSÉ BONIFÁCIO', 'MACAUBAL', 'MACEDÔNIA', 'MAGDA', 'MARAPOAMA', 'MARINÓPOLIS', 'MENDONÇA', 'MERIDIANO', 'MESÓPOLIS', 'MIRA ESTRELA', 'MIRASSOL', 'MIRASSOLÂNDIA', 
'MONÇÕES', 'MONTE APRAZÍVEL', 'NEVES PAULISTA', 'NHANDEARA', 'NIPOÃ', 'NOVA ALIANÇA', 'NOVA CANAÃ PAULISTA', 'NOVA GRANADA', 'NOVAIS', 'NOVO HORIZONTE', 'ONDA VERDE', 'ORINDIÚVA', 'OUROESTE', 'PALESTINA', 'PALMARES PAULISTA', "PALMEIRA D'OESTE", 'PARAÍSO', 'PARANAPUÃ', 'PARISI', 'PAULO DE FARIA', 'PEDRANÓPOLIS', 'PINDORAMA', 'PIRANGI', 'PLANALTO', 'POLONI', 'PONTALINDA', 'PONTES GESTAL', 'POPULINA', 'POTIRENDABA', 'RIOLÂNDIA', 'RUBINÉIA', 'SALES', 'SANTA ADÉLIA', 'SANTA ALBERTINA', "SANTA CLARA D'OESTE", 'SANTA FÉ DO SUL', "SANTA RITA D'OESTE", 'SANTA SALETE', 'SANTANA DA PONTE PENSA', 'SÃO FRANCISCO', 
'SÃO JOÃO DAS DUAS PONTES', 'SÃO JOÃO DE IRACEMA', 'SÃO JOSÉ DO RIO PRETO', 'SEBASTIANÓPOLIS DO SUL', 'TABAPUÃ', 'TANABI', 'TRÊS FRONTEIRAS', 'TURMALINA', 'UBARANA', 'UCHOA', 'UNIÃO PAULISTA', 'URÂNIA', 'URUPÊS', 'VALENTIM GENTIL', 'VITÓRIA BRASIL', 'VOTUPORANGA', 'ZACARIAS']
drsBauru = ['ÁGUAS DE SANTA BÁRBARA', 'AGUDOS', 'ANHEMBI', 'ARANDU', 'AREALVA', 'AREIÓPOLIS', 'AVAÍ', 'AVARÉ', 'BALBINOS', 'BARÃO DE ANTONINA', 'BARIRI', 'BARRA BONITA', 'BAURU', 'BOCAINA', 'BOFETE', 'BORACÉIA', 'BOREBI', 'BOTUCATU', 'BROTAS', 'CABRÁLIA PAULISTA', 'CAFELÂNDIA', 'CERQUEIRA CÉSAR', 'CONCHAS', 'CORONEL MACEDO', 'DOIS CÓRREGOS', 'DUARTINA', 'FARTURA', 'GETULINA', 'GUAIÇARA', 'IACANGA', 'IARAS', 'IGARAÇU DO TIETÊ', 'ITAÍ', 'ITAJU', 'ITAPORANGA', 'ITAPUÍ', 'ITATINGA', 'JAÚ', 'LARANJAL PAULISTA', 'LENÇÓIS PAULISTA', 'LINS', 'LUCIANÓPOLIS', 'MACATUBA', 'MANDURI', 'MINEIROS DO TIETÊ', 'PARANAPANEMA', 'PARDINHO', 'PAULISTÂNIA', 'PEDERNEIRAS', 'PEREIRAS', 'PIRAJU', 'PIRAJUÍ', 'PIRATININGA', 'PONGAÍ', 'PORANGABA', 'PRATÂNIA', 'PRESIDENTE ALVES', 'PROMISSÃO', 'REGINÓPOLIS', 'SABINO', 'SÃO MANUEL', 'SARUTAIÁ', 'TAGUAÍ', 'TAQUARITUBA', 'TEJUPÁ', 'TORRE DE PEDRA', 'TORRINHA', 'URU']
drsMarilia = ['ADAMANTINA', 'ÁLVARO DE CARVALHO', 'ALVINLÂNDIA', 'ARCO-ÍRIS', 'ASSIS', 'BASTOS', 'BERNARDINO DE CAMPOS', 'BORÁ', 'CAMPOS NOVOS PAULISTA', 'CÂNDIDO MOTA', 'CANITAR', 'CHAVANTES', 'CRUZÁLIA', 'ECHAPORÃ', 'ESPÍRITO SANTO DO TURVO', 'FERNÃO', 'FLÓRIDA PAULISTA', 'FLORÍNEA', 'GÁLIA', 'GARÇA', 'GUAIMBÊ', 'GUARANTÃ', 'HERCULÂNDIA', 'IACRI', 'IBIRAREMA', 'INÚBIA PAULISTA', 'IPAUSSU', 'JÚLIO MESQUITA', 'LUCÉLIA', 'LUPÉRCIO', 'LUTÉCIA', 'MARACAÍ', 'MARIÁPOLIS', 'MARÍLIA', 'OCAUÇU', 'ÓLEO', 'ORIENTE', 'OSCAR BRESSANE', 'OSVALDO CRUZ', 'OURINHOS', 'PACAEMBU', 'PALMITAL', 'PARAGUAÇU PAULISTA', 'PARAPUÃ', 'PEDRINHAS PAULISTA', 'PLATINA', 'POMPÉIA', 'PRACINHA', 'QUEIROZ', 'QUINTANA', 'RIBEIRÃO DO SUL', 'RINÓPOLIS', 'SAGRES', 'SALMOURÃO', 'SALTO GRANDE', 'SANTA CRUZ DO RIO PARDO', 'SÃO PEDRO DO TURVO', 'TARUMÃ', 'TIMBURI', 'TUPÃ', 'UBIRAJARA', 'VERA CRUZ'] 
drsSorocaba = ['ALAMBARI', 'ALUMÍNIO', 'ANGATUBA', 'APIAÍ', 'ARAÇARIGUAMA', 'ARAÇOIABA DA SERRA', 'BARRA DO CHAPÉU', 'BOITUVA', 'BOM SUCESSO DE ITARARÉ', 'BURI', 'CAMPINA DO MONTE ALEGRE', 'CAPÃO BONITO', 'CAPELA DO ALTO', 'CERQUILHO', 'CESÁRIO LANGE', 'GUAPIARA', 'GUAREÍ', 'IBIÚNA', 'IPERÓ', 'ITABERÁ', 'ITAOCA', 'ITAPETININGA', 'ITAPEVA', 'ITAPIRAPUÃ PAULISTA', 'ITARARÉ', 'ITU', 'JUMIRIM', 'MAIRINQUE', 'NOVA CAMPINA', 'PIEDADE', 'PILAR DO SUL', 'PORTO FELIZ', 'QUADRA', 'RIBEIRA', 'RIBEIRÃO BRANCO', 'RIBEIRÃO GRANDE', 'RIVERSUL', 'SALTO', 'SALTO DE PIRAPORA', 'SÃO MIGUEL ARCANJO', 'SÃO ROQUE', 'SARAPUÍ', 'SOROCABA', 
'TAPIRAÍ', 'TAQUARIVAÍ', 'TATUÍ', 'TIETÊ', 'VOTORANTIM',
'ORIENTE', 'OSCAR BRESSANE', 'OSVALDO CRUZ', 'OURINHOS', 'PACAEMBU', 'PALMITAL', 'PARAGUAÇU PAULISTA', 'PARAPUÃ', 'PEDRINHAS PAULISTA', 'PLATINA', 'POMPÉIA', 'PRACINHA', 'QUEIROZ', 'QUINTANA', 'RIBEIRÃO DO SUL', 'RINÓPOLIS', 'SAGRES', 'SALMOURÃO', 'SALTO GRANDE', 'SANTA CRUZ DO RIO PARDO', 'SÃO PEDRO DO TURVO', 'TARUMÃ', 'TIMBURI', 'TUPÃ', 'UBIRAJARA', 'VERA CRUZ']
drsPresidentePrudente = ['ALFREDO MARCONDES', 'ÁLVARES MACHADO', 'ANHUMAS', 'CAIABU', 'CAIUÁ', 'DRACENA', 'EMILIANÓPOLIS', 'ESTRELA DO NORTE', 'EUCLIDES DA CUNHA PAULISTA', 'FLORA RICA', 'IEPÊ', 'INDIANA', 'IRAPURU', 'JOÃO RAMALHO', 'JUNQUEIRÓPOLIS', 'MARABÁ PAULISTA', 'MARTINÓPOLIS', 'MIRANTE DO PARANAPANEMA', 'MONTE CASTELO', 'NANTES', 'NARANDIBA', 'NOVA GUATAPORANGA', 'OURO VERDE', 'PANORAMA', 'PAULICÉIA', 'PIQUEROBI', 'PIRAPOZINHO', 'PRESIDENTE BERNARDES', 'PRESIDENTE EPITÁCIO', 'PRESIDENTE PRUDENTE', 'PRESIDENTE VENCESLAU', 'QUATÁ', 'RANCHARIA', 'REGENTE FEIJÓ', 'RIBEIRÃO DOS ÍNDIOS', 'ROSANA', 'SANDOVALINA', 'SANTA MERCEDES', 'SANTO ANASTÁCIO', 'SANTO EXPEDITO', "SÃO JOÃO DO PAU D'ALHO", 'TACIBA', 'TARABAI', 'TEODORO SAMPAIO', 'TUPI PAULISTA']
drsCampinas = ['ÁGUAS DE LINDÓIA', 'AMERICANA', 'AMPARO', 'ARTUR NOGUEIRA', 'ATIBAIA', 'BOM JESUS DOS PERDÕES', 'BRAGANÇA PAULISTA', 'CABREÚVA', 'CAMPINAS', 'CAMPO LIMPO PAULISTA', 'COSMÓPOLIS', 'HOLAMBRA', 'HORTOLÂNDIA', 'INDAIATUBA', 'ITATIBA', 'ITUPEVA', 'JAGUARIÚNA', 'JARINU', 'JOANÓPOLIS', 'JUNDIAÍ', 'LINDÓIA', 'LOUVEIRA', 'MONTE ALEGRE DO SUL', 'MONTE MOR', 'MORUNGABA', 'NAZARÉ PAULISTA', 'NOVA ODESSA', 'PAULÍNIA', 'PEDRA BELA', 'PEDREIRA', 'PINHALZINHO', 'PIRACAIA', "SANTA BÁRBARA D'OESTE", 'SANTO ANTÔNIO DA POSSE', 'SERRA NEGRA', 'SOCORRO', 'SUMARÉ', 'TUIUTI', 'VALINHOS', 'VARGEM', 'VÁRZEA PAULISTA', 'VINHEDO']
drsAracatuba = ['ALTO ALEGRE', 'ANDRADINA', 'ARAÇATUBA', 'AURIFLAMA', 'AVANHANDAVA', 'BARBOSA', 'BENTO DE ABREU', 'BILAC', 'BIRIGUI', 'BRAÚNA', 'BREJO ALEGRE', 'BURITAMA', 'CASTILHO', 'CLEMENTINA', 'COROADOS', 'GABRIEL MONTEIRO', 'GLICÉRIO', 'GUARAÇAÍ', 'GUARARAPES', 'GUZOLÂNDIA', 'ILHA SOLTEIRA', 'ITAPURA', 'LAVÍNIA', 'LOURDES', 'LUIZIÂNIA', 'MIRANDÓPOLIS', 'MURUTINGA DO SUL', 'NOVA CASTILHO', 'NOVA INDEPENDÊNCIA', 'NOVA LUZITÂNIA', 'PENÁPOLIS', 'PEREIRA BARRETO', 'PIACATU', 'RUBIÁCEA', 'SANTO ANTÔNIO DO ARACANGUÁ', 'SANTÓPOLIS DO AGUAPEÍ', 'SUD MENNUCCI', 'SUZANÁPOLIS', 'TURIÚBA', 'VALPARAÍSO']
drsSaoPaulo = ['ARUJÁ', 'BARUERI', 'BIRITIBA MIRIM', 'CAIEIRAS', 'CAJAMAR', 'CARAPICUÍBA', 'COTIA', 'DIADEMA', 'EMBU DAS ARTES', 'EMBU-GUAÇU', 'FERRAZ DE VASCONCELOS', 'FRANCISCO MORATO', 'FRANCO DA ROCHA', 'GUARAREMA', 'GUARULHOS', 'ITAPECERICA DA SERRA', 'ITAPEVI', 'ITAQUAQUECETUBA', 'JANDIRA', 'JUQUITIBA', 'MAIRIPORÃ', 'MAUÁ', 'MOGI DAS CRUZES', 'OSASCO', 'PIRAPORA DO BOM JESUS', 'POÁ', 'RIBEIRÃO PIRES', 'RIO GRANDE DA SERRA', 'SALESÓPOLIS', 'SANTA ISABEL', 'SANTANA DE PARNAÍBA', 'SANTO ANDRÉ', 'SÃO BERNARDO DO CAMPO', 'SÃO CAETANO DO SUL', 'SÃO LOURENÇO DA SERRA', 'SÃO PAULO', 'SUZANO', 'TABOÃO DA SERRA', 'VARGEM GRANDE PAULISTA']
drsTaubate = ['APARECIDA', 'ARAPEÍ', 'AREIAS', 'BANANAL', 'CAÇAPAVA', 'CACHOEIRA PAULISTA', 'CAMPOS DO JORDÃO', 'CANAS', 'CARAGUATATUBA', 'CRUZEIRO', 'CUNHA', 'GUARATINGUETÁ', 'IGARATÁ', 'ILHABELA', 'JACAREÍ', 'JAMBEIRO', 'LAGOINHA', 'LAVRINHAS', 'LORENA', 'MONTEIRO LOBATO', 'NATIVIDADE DA SERRA', 'PARAIBUNA', 'PINDAMONHANGABA', 'PIQUETE', 'POTIM', 'QUELUZ', 'REDENÇÃO DA SERRA', 'ROSEIRA', 'SANTA BRANCA', 'SANTO ANTÔNIO DO PINHAL', 'SÃO BENTO DO SAPUCAÍ', 'SÃO JOSÉ DO BARREIRO', 'SÃO JOSÉ DOS CAMPOS', 'SÃO LUIZ DO PARAITINGA', 'SÃO SEBASTIÃO', 'SILVEIRAS', 'TAUBATÉ', 'TREMEMBÉ', 'UBATUBA']
drsPiracicaba = ['ÁGUAS DE SÃO PEDRO', 'ANALÂNDIA', 'ARARAS', 'CAPIVARI', 'CHARQUEADA', 'CONCHAL', 'CORDEIRÓPOLIS', 'CORUMBATAÍ', 'ELIAS FAUSTO', 'ENGENHEIRO COELHO', 'IPEÚNA', 'IRACEMÁPOLIS', 'ITIRAPINA', 'LEME', 'LIMEIRA', 'MOMBUCA', 'PIRACICABA', 'PIRASSUNUNGA', 'RAFARD', 'RIO CLARO', 'RIO DAS PEDRAS', 'SALTINHO', 'SANTA CRUZ DA CONCEIÇÃO', 'SANTA GERTRUDES', 'SANTA MARIA DA SERRA', 'SÃO PEDRO']
drsRibeiraoPreto = ['ALTINÓPOLIS', 'BARRINHA', 'BATATAIS', 'BRODOWSKI', 'CAJURU', 'CÁSSIA DOS COQUEIROS', 'CRAVINHOS', 'DUMONT', 'GUARIBA', 'GUATAPARÁ', 'JABOTICABAL', 'JARDINÓPOLIS', 'LUÍS ANTÔNIO', 'MONTE ALTO', 'PITANGUEIRAS', 'PONTAL', 'PRADÓPOLIS', 'RIBEIRÃO PRETO', 'SANTA CRUZ DA ESPERANÇA', 'SANTA RITA DO PASSA QUATRO', 'SANTA ROSA DE VITERBO', 'SANTO ANTÔNIO DA ALEGRIA', 'SÃO SIMÃO', 'SERRA AZUL', 'SERRANA', 'SERTÃOZINHO']
drsAraraquara = ['AMÉRICO BRASILIENSE', 'ARARAQUARA', 'BOA ESPERANÇA DO SUL', 'BORBOREMA', 'CÂNDIDO RODRIGUES', 'DESCALVADO', 'DOBRADA', 'DOURADO', 'GAVIÃO PEIXOTO', 'IBATÉ', 
'IBITINGA', 'ITÁPOLIS', 'MATÃO', 'MOTUCA', 'NOVA EUROPA', 'PORTO FERREIRA', 'RIBEIRÃO BONITO', 'RINCÃO', 'SANTA ERNESTINA', 'SANTA LÚCIA', 'SÃO CARLOS', 'TABATINGA', 'TAQUARITINGA', 'TRABIJU']
drsFranca = ['ARAMINA', 'BURITIZAL', 'CRISTAIS PAULISTA', 'FRANCA', 'GUARÁ', 'IGARAPAVA', 'IPUÃ', 'ITIRAPUÃ', 'ITUVERAVA', 'JERIQUARA', 'MIGUELÓPOLIS', 'MORRO AGUDO', 'NUPORANGA', 'ORLÂNDIA', 'PATROCÍNIO PAULISTA', 'PEDREGULHO', 'RESTINGA', 'RIBEIRÃO CORRENTE', 'RIFAINA', 'SALES OLIVEIRA', 'SÃO JOAQUIM DA BARRA', 'SÃO JOSÉ DA BELA VISTA']
drsSaoJoaoDaBoaVista = ['AGUAÍ', 'ÁGUAS DA PRATA', 'CACONDE', 'CASA BRANCA', 'DIVINOLÂNDIA', 'ESPÍRITO SANTO DO PINHAL', 'ESTIVA GERBI', 'ITAPIRA', 'ITOBI', 'MOCOCA', 'MOGI GUAÇU', 'MOGI MIRIM', 'SANTA CRUZ DAS PALMEIRAS', 'SANTO ANTÔNIO DO JARDIM', 'SÃO JOÃO DA BOA VISTA', 'SÃO JOSÉ DO RIO PARDO', 'SÃO SEBASTIÃO DA GRAMA', 'TAMBAÚ', 'TAPIRATIBA', 'VARGEM GRANDE DO SUL']
drsBarretos = ['ALTAIR', 'BARRETOS', 'BEBEDOURO', 'CAJOBI', 'COLINA', 'COLÔMBIA', 'GUAÍRA', 'GUARACI', 'JABORANDI', 'MONTE AZUL PAULISTA', 'OLÍMPIA', 'SEVERÍNIA', 'TAIAÇU', 
'TAIÚVA', 'TAQUARAL', 'TERRA ROXA', 'VIRADOURO', 'VISTA ALEGRE DO ALTO']
drsRegistro = ['BARRA DO TURVO', 'CAJATI', 'CANANÉIA', 'ELDORADO', 'IGUAPE', 'ILHA COMPRIDA', 'IPORANGA', 'ITARIRI', 'JACUPIRANGA', 'JUQUIÁ', 'MIRACATU', 'PARIQUERA-AÇU', 'PEDRO DE TOLEDO', 'REGISTRO', 'SETE BARRAS']

todasAsDrs = ['Grande São Paulo','Araçatuba','Araraquara','Baixada','Barretos','Bauru','Campinas','Franca','Marília','Piracicaba','Presidente Prudente','Registro','Ribeirão Preto','São João Da Boa Vista','Rio Preto','Sorocaba','Taubaté','Estado']
# distribuindo as DRSes na planilha
def distribuirDrsNaPlanilha():
    linha = 2

    while planilha["C"+str(linha)].value != 'Ignorado':
        # print("C"+str(linha))
        cidade = planilha["C"+str(linha)].value.upper()

        if cidade in drsRioPreto:
            planilha["B"+str(linha)].value = todasAsDrs[14]
        elif cidade in drsBauru:
            planilha["B"+str(linha)].value = todasAsDrs[5]
        elif cidade in drsMarilia:
            planilha["B"+str(linha)].value = todasAsDrs[8]
        elif cidade in drsSorocaba:
            planilha["B"+str(linha)].value = todasAsDrs[15]
        elif cidade in drsPresidentePrudente:
            planilha["B"+str(linha)].value = todasAsDrs[10]
        elif cidade in drsCampinas:
            planilha["B"+str(linha)].value = todasAsDrs[6]
        elif cidade in drsAracatuba:
            planilha["B"+str(linha)].value = todasAsDrs[1]
        elif cidade in drsSaoPaulo:
            planilha["B"+str(linha)].value = todasAsDrs[0]
        elif cidade in drsTaubate:
            planilha["B"+str(linha)].value = todasAsDrs[16]
        elif cidade in drsPiracicaba:
            planilha["B"+str(linha)].value = todasAsDrs[9]
        elif cidade in drsRibeiraoPreto:
            planilha["B"+str(linha)].value = todasAsDrs[12]
        elif cidade in drsAraraquara:
            planilha["B"+str(linha)].value = todasAsDrs[2]
        elif cidade in drsFranca:
            planilha["B"+str(linha)].value = todasAsDrs[7]
        elif cidade in drsSaoJoaoDaBoaVista:
            planilha["B"+str(linha)].value = todasAsDrs[13]
        elif cidade in drsBarretos:
            planilha["B"+str(linha)].value = todasAsDrs[4]
        elif cidade in drsRegistro:
            planilha["B"+str(linha)].value = todasAsDrs[11]
        else:
            planilha["B"+str(linha)].value = todasAsDrs[3]
        
        linha += 1

distribuirDrsNaPlanilha()

""" Criando uma nova planilha com focada nas DRSes """
arquivo.create_sheet('PLANILHA_DRS')
planilha2 = arquivo["PLANILHA_DRS"]

todasAsColunas = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY"]
def criarEstruturaDaPlanilhaDRS():
    for i in range(0,49,3):
        planilha2[todasAsColunas[i]+"1"].value = todasAsDrs[round(i/3)]
        planilha2.merge_cells(todasAsColunas[i]+"1:"+todasAsColunas[i+2]+"1")

        planilha2[todasAsColunas[i]+"2"].value = 'Cidade'
        planilha2[todasAsColunas[i+1]+"2"].value = 'Casos'
        planilha2[todasAsColunas[i+2]+"2"].value = 'Mortes'
criarEstruturaDaPlanilhaDRS()

# função que soma todos os itens da coluna
def somarValoresDaColuna(planilhaEscolhida,coluna,linha,ultimaLinha):
    array = []
    while linha < ultimaLinha:
        array.append(planilhaEscolhida[coluna+str(linha)].value)
        
        linha += 1

    return functools.reduce(lambda a, b: int(a)+int(b), array)

todasAsMedias = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
def distribuirCidadesNasDRS():
    linha = 2
    posLinhasVazias = [3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3]

    while planilha["C"+str(linha)].value != 'Ignorado':
        drs = planilha["B"+str(linha)].value
        cidade = planilha["C"+str(linha)].value
        numCasos = planilha["D"+str(linha)].value
        numMortes = planilha["E"+str(linha)].value

        for i in range(0,49,3):
            if drs == planilha2[todasAsColunas[i]+"1"].value:
                planilha2[todasAsColunas[i]+str(posLinhasVazias[round(i/3)])].value = cidade
                planilha2[todasAsColunas[i+1]+str(posLinhasVazias[round(i/3)])].value = numCasos
                planilha2[todasAsColunas[i+2]+str(posLinhasVazias[round(i/3)])].value = numMortes
                
                posLinhasVazias[round(i/3)] += 1

        linha += 1

    for c in range(1,50,3):
        linhaVazia = posLinhasVazias[round(c/3)]

        planilha2[todasAsColunas[c-1]+str(linhaVazia)].value = 'TOTAL:'
        planilha2[todasAsColunas[c]+str(linhaVazia)].value = somarValoresDaColuna(planilha2,todasAsColunas[c],3,linhaVazia)
        planilha2[todasAsColunas[c+1]+str(linhaVazia)].value = somarValoresDaColuna(planilha2,todasAsColunas[c+1],3,linhaVazia)

        todasAsMedias[round(c/3)] = round(int(planilha2[todasAsColunas[c+1]+str(linhaVazia)].value)*100/int(planilha2[todasAsColunas[c]+str(linhaVazia)].value), 2)
distribuirCidadesNasDRS()

mediaSP = round(somarValoresDaColuna(planilha,todasAsColunas[4],2,planilha.max_row-2)*100/somarValoresDaColuna(planilha,todasAsColunas[3],2,planilha.max_row-2),2)
todasAsMedias[17] = mediaSP

# colocando as informações em gráfico e salvando -o
plt.barh(todasAsDrs,todasAsMedias,color=['lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','lightblue','#ff4137'])
plt.grid(color='#ccc', axis='x')

plt.title('Taxa De Letalidade De Cada Departamento Regional De Saúde',fontsize=18)
plt.ylabel('DEPARTAMENTO', fontsize=14)
plt.xlabel('LETALIDADE (%)', fontsize=14)

plt.savefig('departamentos.png',orientation='landscape',bbox_inches='tight')

# executando automaticamente o outro script
exec(open('relatorio.py').read())

# salvando o arquivo xlsx
arquivo.save('municipios_drs.xlsx')
print('ARQUIVO SALVO COM SUCESSO!')